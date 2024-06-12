# main.py
import pandas as pd
import openpyxl

# Define function to convert column alphabets to numerical indices
def alphabet_to_number(col_idx):
    col_idx = col_idx.upper()
    num = 0
    for letter in col_idx:
        num = num * 26 + (ord(letter) - ord('A')) + 1
    return num - 1  # Subtract 1 to start from 0


# alphabet_to_number = {
#     'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5, 'G': 6, 'H': 7, 'I': 8, 'J': 9,
#     'K': 10, 'L': 11, 'M': 12, 'N': 13, 'O': 14, 'P': 15, 'Q': 16, 'R': 17, 'S': 18,
#     'T': 19, 'U': 20, 'V': 21, 'W': 22, 'X': 23, 'Y': 24, 'Z': 25, 'AA': 26, 'AB': 27,
#     'AC': 28, 'AD': 29, 'AE': 30, 'AF': 31, 'AG': 32, 'AH': 33, 'AI': 34, 'AJ': 35,
#     'AK': 36, 'AL': 37, 'AM': 38, 'AN': 39, 'AO': 40, 'AP': 41, 'AQ': 42, 'AR': 43,
#     'AS': 44, 'AT': 45, 'AU': 46, 'AV': 47, 'AW': 48, 'AX': 49, 'AY': 50, 'AZ': 51,
#     'BA': 52, 'BB': 53, 'BC': 54, 'BD': 55, 'BE': 56, 'BF': 57, 'BG': 58, 'BH': 59,
#     'BI': 60, 'BJ': 61, 'BK': 62, 'BL': 63, 'BM': 64, 'BN': 65, 'BO': 66
# }


# # Function to get cell values from an Excel file

# import pandas as pd

# # Mapping of alphabets to their respective numerical values
# def alphabet_to_number(col_idx):
#     col_idx = col_idx.upper()
#     num = 0
#     for letter in col_idx:
#         num = num * 26 + (ord(letter) - ord('A')) + 1
#     return num - 1  # Subtract 1 to start from 0


# Define function to get cell values from an Excel file
def get_cell_value(file_path, sheet_name, cell_ref):
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        val = {}

        for cell in cell_ref:
            col_idx = ''.join(filter(str.isalpha, cell))
            row_idx = int(''.join(filter(str.isdigit, cell))) - 1

            if col_idx.isalpha() and row_idx >= 0:
                col_num = alphabet_to_number(col_idx)
                val[cell] = df.iloc[row_idx, col_num]
            else:
                print(f"Invalid cell reference '{cell}'")

        return val

    except Exception as e:
        print(f"Error occurred while reading cell {cell_ref} from sheet {sheet_name} in {file_path}: {e}")
        return None

# Define function to write cell values to Excel
def write_to_excel(file1_cells, file2_cells, file2_cellsV1, file2_cellsV2, file1_cellsV1, file1_cellsV2, filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Comparison"

    sheet['A1'] = "cell_value_file1"
    sheet['B1'] = "cell_value_file2"

    # row_idx = 2
    # for cell, value in file1_cells.items():
    #     sheet.cell(row=row_idx, column=1).value = value
    #     row_idx += 1

    # row_idx = 2
    # for cell, value in file2_cells.items():
    #     sheet.cell(row=row_idx, column=2).value = value
    #     row_idx += 1

    # row_idx = len(file1_cells) + 2
    # for cell, value in file2_cellsV1.items():
    #     sheet.cell(row=row_idx, column=2).value = value
    #     row_idx += 1

    # row_idx = len(file2_cellsV1) + 2
    # for cell, value in file2_cellsV2.items():
    #     sheet.cell(row=row_idx, column=2).value = value
    #     row_idx += 1

    # row_idx = len(file2_cellsV2) + 2
    # for cell, value in file1_cellsV1.items():
    #     sheet.cell(row=row_idx, column=1).value = value
    #     row_idx += 1

    # row_idx = len(file1_cellsV1) + 2
    # for cell, value in file1_cellsV2.items():
    #     sheet.cell(row=row_idx, column=1).value = value
    #     row_idx += 1

    # workbook.save(filename)

    row_idx = 2
    for cell, value in file1_cells.items():
        sheet.cell(row=row_idx, column=1).value = value
        row_idx += 1

    row_idx = 2
    for cell, value in file2_cells.items():
        sheet.cell(row=row_idx, column=2).value = value
        row_idx += 1

    row_idx = len(file1_cells) + 2
    for cell, value in file2_cellsV1.items():
        sheet.cell(row=row_idx, column=1).value = value
        row_idx += 1

    row_idx = len(file2_cells) + 2
    for cell, value in file2_cellsV2.items():
        sheet.cell(row=row_idx, column=2).value = value
        row_idx += 1

    row_idx = len(file1_cells) + len(file2_cellsV1) + 2
    for cell, value in file1_cellsV1.items():
        sheet.cell(row=row_idx, column=1).value = value
        row_idx += 1

    row_idx = len(file2_cells)  + len(file2_cellsV2) + 2
    for cell, value in file1_cellsV2.items():
        sheet.cell(row=row_idx, column=2).value = value
        row_idx += 1

    workbook.save(filename)
